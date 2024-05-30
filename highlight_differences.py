#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Define the file name
compare_file = 'production_and_qa_code_compare.xlsx'

# Load the workbook
wb = load_workbook(compare_file)

# Define tabs to iterate over
tabs = ['master_file_compare', 'join_file_compare', 'detail_file_compare']

# Define a fill color for highlighting differences
diff_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Function to apply highlighting to cell differences in a sheet
def highlight_cell_differences(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    prod_columns = []
    qa_columns = []

    # Identify columns for Production and QA values
    for i, header in enumerate(header_row, 1):
        if '_prod' in header:
            prod_columns.append(i)
        elif '_qa' in header:
            qa_columns.append(i)

    # Iterate over rows and highlight differences
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column, values_only=False):
        for prod_col_index, qa_col_index in zip(prod_columns, qa_columns):
            prod_cell = row[prod_col_index-1]  # Adjusting index for zero-based index
            qa_cell = row[qa_col_index-1]  # Adjusting index for zero-based index

            if prod_cell.value != qa_cell.value:
                prod_cell.fill = diff_fill
                qa_cell.fill = diff_fill

# Iterate over each comparison tab and apply highlighting
for tab_name in tabs:
    if tab_name in wb.sheetnames:
        sheet = wb[tab_name]
        highlight_cell_differences(sheet)
    else:
        print(f"Tab {tab_name} not found in the workbook.")

# Save the workbook with highlighted differences
wb.save('highlighted_differences_' + compare_file)

print("Cells with differences in the Excel file have been highlighted.")


# In[ ]:




